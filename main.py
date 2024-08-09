import os
import tkinter as tk
from tkinter import filedialog, messagebox
import cv2
from PIL import Image, ImageTk
import pytesseract
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime

import util


class App:
    def __init__(self):
        self.main_window = tk.Tk()
        self.main_window.geometry("1200x520+350+100")

        self.product_options = ["S1125", "DIPL", "Spl55", "Spl90", "2040", "2045", "2055"]
        self.weight_ranges = {
            "S1125": "Weight range: 0.1-0.20 kg",
            "DIPL": "Weight range: 0.5-0.6 kg",
            "Spl55": "Weight range: 0.1-0.20 kg",
            "Spl90": "Weight range: 1.01-1.05 kg",
            "2040": "Weight range: 1.1-1.4 kg",
            "2045": "Weight range: 1.7-2.0 kg",
            "2055": "Weight range: 2.6-2.9 kg"
        }

        self.db_dir = './db'
        if not os.path.exists(self.db_dir):
            os.mkdir(self.db_dir)

        self.setup_capture_upload_page()

    def setup_capture_upload_page(self):
        self.capture_image_button = util.get_button(self.main_window, 'Capture Image', 'gray', self.capture_image, fg='black')
        self.capture_image_button.place(x=750, y=300)

        self.upload_image_button = util.get_button(self.main_window, 'Upload Image', 'gray', self.upload_image, fg='black')
        self.upload_image_button.place(x=750, y=400)

        self.webcam_label = util.get_img_label(self.main_window)
        self.webcam_label.place(x=10, y=0, width=700, height=500)

        self.add_webcam(self.webcam_label)

    def add_webcam(self, label):
        if 'cap' not in self.__dict__:
            self.cap = cv2.VideoCapture(0)

        self._label = label
        self.process_webcam()

    def process_webcam(self):
        ret, frame = self.cap.read()

        self.most_recent_capture_arr = frame
        img_ = cv2.cvtColor(self.most_recent_capture_arr, cv2.COLOR_BGR2RGB)
        self.most_recent_capture_pil = Image.fromarray(img_)
        imgtk = ImageTk.PhotoImage(image=self.most_recent_capture_pil)
        self._label.imgtk = imgtk
        self._label.configure(image=imgtk)

        self._label.after(20, self.process_webcam)

    def capture_image(self):
        self.capture_image_window = tk.Toplevel(self.main_window)
        self.capture_image_window.geometry("1200x520+370+120")

        self.accept_button_capture_image = util.get_button(self.capture_image_window, 'Accept', 'green', self.accept_capture_image)
        self.accept_button_capture_image.place(x=750, y=300)

        self.try_again_button_capture_image = util.get_button(self.capture_image_window, 'Try again', 'red', self.try_again_capture_image)
        self.try_again_button_capture_image.place(x=750, y=400)

        self.capture_label = util.get_img_label(self.capture_image_window)
        self.capture_label.place(x=10, y=0, width=700, height=500)

        self.add_img_to_label(self.capture_label)

    def upload_image(self):
        file_path = filedialog.askopenfilename()
        if file_path:
            self.most_recent_capture_pil = Image.open(file_path)
            self.most_recent_capture_arr = cv2.cvtColor(cv2.imread(file_path), cv2.COLOR_BGR2RGB)

            self.upload_image_window = tk.Toplevel(self.main_window)
            self.upload_image_window.geometry("1200x520+370+120")

            self.accept_button_upload_image = util.get_button(self.upload_image_window, 'Accept', 'green', lambda: self.accept_upload_image(file_path))
            self.accept_button_upload_image.place(x=750, y=300)

            self.try_again_button_upload_image = util.get_button(self.upload_image_window, 'Try again', 'red', self.try_again_upload_image)
            self.try_again_button_upload_image.place(x=750, y=400)

            self.upload_label = util.get_img_label(self.upload_image_window)
            self.upload_label.place(x=10, y=0, width=700, height=500)

            self.add_img_to_label(self.upload_label)

    def try_again_capture_image(self):
        self.capture_image_window.destroy()

    def try_again_upload_image(self):
        self.upload_image_window.destroy()

    def add_img_to_label(self, label):
        imgtk = ImageTk.PhotoImage(image=self.most_recent_capture_pil)
        label.imgtk = imgtk
        label.configure(image=imgtk)

        self.capture_image_capture = self.most_recent_capture_arr.copy()

    def accept_capture_image(self):
        temp_image_path = os.path.join(self.db_dir, 'temp_capture_image.jpg')
        cv2.imwrite(temp_image_path, self.capture_image_capture)
        self.run_function_with_path(temp_image_path)

    def accept_upload_image(self, file_path):
        self.run_function_with_path(file_path)

    def run_function_with_path(self, file_path):
        detected_text = extract_and_detect(file_path, self.db_dir)
        self.setup_product_selection_page(detected_text)

    def setup_product_selection_page(self, detected_text):
        self.main_window.destroy()  # Close the previous window
        self.product_window = tk.Tk()
        self.product_window.geometry("1200x520+350+100")

        self.detected_text_label = util.get_text_label(self.product_window, f"Detected Text: {detected_text}")
        self.detected_text_label.place(x=750, y=50)

        self.selected_product = tk.StringVar(self.product_window)
        self.selected_product.set(self.product_options[0])  # default value

        self.product_dropdown = tk.OptionMenu(self.product_window, self.selected_product, *self.product_options, command=self.update_weight_range)
        self.product_dropdown.place(x=750, y=150)

        self.text_label_product_window = util.get_text_label(self.product_window, 'Please, select a product:')
        self.text_label_product_window.place(x=750, y=100)

        self.weight_range_label = util.get_text_label(self.product_window, self.weight_ranges[self.selected_product.get()])
        self.weight_range_label.place(x=750, y=200)

        self.submit_button_product_window = util.get_button(self.product_window, 'Submit', 'gray', lambda: self.save_to_excel(detected_text), fg='black')
        self.submit_button_product_window.place(x=750, y=250)

        self.retry_button_product_window = util.get_button(self.product_window, 'Retry', 'gray', self.restart_app, fg='black')
        self.retry_button_product_window.place(x=750, y=350)

    def update_weight_range(self, selection):
        self.weight_range_label.config(text=self.weight_ranges[selection])

    def save_to_excel(self, detected_text):
        try:
            detected_weight = float(detected_text)
            product = self.selected_product.get()

            if product in self.weight_ranges:
                weight_range_str = self.weight_ranges[product]
                weight_range = weight_range_str.split(': ')[1].split(' ')[0].split('-')

                if not (float(weight_range[0]) <= detected_weight <= float(weight_range[1])):
                    raise ValueError(
                        f"Detected weight ({detected_weight} kg) is not within the specified range ({weight_range_str})")

            excel_file_path = os.path.join(self.db_dir, 'detected_values.xlsx')

            if not os.path.exists(excel_file_path):
                wb = Workbook()
                ws = wb.active
                ws.title = "Detected Values"
                ws.append(["Timestamp", "Detected Text", "Product Name", "Weight Range"])
                wb.save(excel_file_path)

            wb = load_workbook(excel_file_path)
            ws = wb["Detected Values"]

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            ws.append([timestamp, float(detected_text), product, self.weight_ranges[product]])

            wb.save(excel_file_path)

            util.msg_box('Success!', 'Data was saved successfully!')

            self.product_window.destroy()

        except ValueError as e:
            util.msg_box('Error!', str(e))
            print(f"Error: {e}")

    def restart_app(self):
        self.product_window.destroy()
        self.__init__()
        self.start()

    def start(self):
        self.main_window.mainloop()


def extract_and_detect(img_path, db_path):
    # Load the image
    image = cv2.imread(img_path)

    if image is None:
        print(f"Error: Unable to load image at {img_path}")
        return "Error: Unable to load image"

    # Switch case-like structure for coordinates based on image name
    if 'Compact' in img_path:
        x, y, w, h = 100, 200, 55, 27  # Example coordinates compact
    elif 'DIPL' in img_path:
        x, y, w, h = 163, 143, 60, 27  # Example coordinates DIPL
    elif 'Spl100' in img_path:
        x, y, w, h = 163, 200, 71, 27  # Example coordinates Spl100
    elif 'temp' in img_path:
        x, y, w, h = 232, 272, 164, 61
    else:
        print("Error: Image name does not match any known coordinates")
        return "Error: Image name does not match any known coordinates"

    # Ensure coordinates are within the image bounds
    height, width, _ = image.shape
    if x < 0 or y < 0 or x + w > width or y + h > height:
        print("Error: ROI coordinates are out of image bounds")
        return "Error: ROI coordinates are out of image bounds"

    # Crop the ROI from the image
    roi = image[y:y+h, x:x+w]

    # Save the cropped ROI as an image
    output_path = os.path.join(db_path, 'extracted_roi_manual.png')
    cv2.imwrite(output_path, roi)
    print(f"Cropped ROI saved to {output_path}")

    # Read the cropped ROI image
    image = cv2.imread(output_path)
    image = cv2.resize(image, None, fx=1.5, fy=1.5)

    # Convert to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Apply Gaussian blur to reduce noise
    blurred = cv2.GaussianBlur(gray, (3, 3), 0)

    # Apply Otsu's thresholding
    _, otsu_thresh = cv2.threshold(blurred, 0, 255, cv2.THRESH_OTSU)

    # Optional: Increase the contrast of the image
    contrast_enhanced = cv2.convertScaleAbs(otsu_thresh, alpha=1.5, beta=0)

    # Morphological operations to enhance the threshold image
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
    morphed = cv2.morphologyEx(contrast_enhanced, cv2.MORPH_OPEN, kernel)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
    morphed = cv2.erode(morphed, kernel)
    cv2.imshow("Processed Image", morphed)
    cv2.waitKey(0)
    cv2.destroyAllWindows()

    # Detect digits from the preprocessed image
    config = '--psm 6 -c tessedit_char_whitelist="0123456789."'
    detected_text = pytesseract.image_to_string(morphed, config=config, lang='lets').replace('\f', '').replace('\n', '')

    # Process the detected text to add a decimal point
    detected_text = ''.join(detected_text.split())
    return detected_text


if __name__ == "__main__":
    app = App()
    app.start()
